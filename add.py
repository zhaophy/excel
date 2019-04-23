#coding:utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb = load_workbook('F:\\excel\\ab.xlsx')
def if_exist(file):
    try:
        wb = load_workbook(file)
    except (IOError),e:
        print e
if_exist("F:\\excel\\ab.xlsx")



