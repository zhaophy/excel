#!/usr/bin/env python
# -*- coding: utf-8 -*-
import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
import string
from openpyxl import Workbook
import operator

letters = string.uppercase
data_a = [{'name':'echo_one','age':'56','rank':'A','Class':'43'},{'age':'47','rank':'A','Class':'9'},{'age':'44','rank':'B+','Class':'9','name':'echo_four','phone':'10086','add':'henan'}]
list = data_a
print list
for i in list:
    print i
keys = list[0].keys()
print keys
r = openpyxl.load_workbook('F:\excel\ok.xlsx')
ws = r.get_sheet_by_name(str('ok'))


for i in range(1, len(list[0]) + 1):
    ws['%s%s' % (letters[i - 1], 1)] = list[0].keys()[i - 1]  #先增加首行
for i in range(1, len(list[0]) + 1):
    for j in range(1, len(list) + 1):
        #if  list[j - 1][keys]
        for m in list[j - 1].keys():
            #print m
            if m not in keys:
                ws['%s%s' % (letters[ws.max_column ], 1)] = m  #如果列表的字典有的key不在首行
                keys.append(m)
                print keys
                ws['%s%s' % (letters[ws.max_column - 1], j + 1)] = list[j - 1][m]
                try :


                    ws['%s%s' % (letters[i - 1], j + 1)] = list[j - 1][keys[i - 1]] #如果列表的字典缺少首行含有的key
                except:
                    ws['%s%s' % (letters[i - 1], j + 1)] = None


            else:
                try:
                    ws['%s%s' % (letters[i - 1], j + 1)] = list[j - 1][keys[i - 1]]
                except:
                    ws['%s%s' % (letters[i - 1], j + 1)] = None

        #except:
           # key
r.save('F:\excel\ok.xlsx')