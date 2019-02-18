#!/usr/bin/env python
# -*- coding: utf-8 -*-
import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
import string

data_a = [{'name':'echo_one','age':'56','rank':'A','Class':'9'},{'name':'echo_two','age':'47','rank':'A','Class':'9'},{'name':'echo_three','age':'44','rank':'B+','Class':'9'}]
print data_a[0].values()
print data_a[0].keys()
print data_a[0].values()[ 2 - 1 ]
letters = string.uppercase
data_b = [{'base':'sword','cat':'cat-5','level':'11'},{'base':'blade','cat':'cat-3','level':'15'},{'base':'shield','cat':'cat-11','level':'331'}]



def excel_ins(file_path,sheet,list):
    r = openpyxl.load_workbook(filename=file_path)
    #ws = r.active
    ws = r.get_sheet_by_name(str(sheet))
    for i in range(1, len(list[0]) + 1):
        # print letters[i]工作表1
        # print "mn"i
        ws['%s%s' % (letters[i - 1], 1)] = list[0].keys()[i - 1]  # 首行插入参数
        for j in range(1, len(list) + 1):
            ws['%s%s' % (letters[i - 1], j + 1)] = list[j - 1].values()[i - 1] #其余行插入数据
    r.save(file_path)
excel_ins('/Users/zhaophy/untitled2/ok.xlsx','ok',data_a)


