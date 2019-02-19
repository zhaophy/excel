#!/usr/bin/env python
# -*- coding: utf-8 -*-
import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
import string
from openpyxl import Workbook
import operator
data_a = [{'name':'echo_one','age':'56','rank':'A','Class':'9'},{'name':'echo_two','age':'47','rank':'A','Class':'9'},{'age':'44','rank':'B+','Class':'9'}]
#print data_a.items()
letters = string.uppercase
for i in data_a:

    d = sorted(i.items(), key=lambda k: k[1])
    print d

def norm_list(list):
    new_list = []
    for i in list:
        d = sorted(i.items(), key=lambda k: k[0])
        new_list.append(d)
    return new_list


def excel_ins(file_path,sheet,list):
    r = openpyxl.load_workbook(filename=file_path)
    #ws = r.active
    ws = r.get_sheet_by_name(str(sheet))
    for i in range(1, len(list[0]) + 1):
        # print letters[i]工作表1
        # print "mn"i
        #print list[0][2][0]
        ws['%s%s' % (letters[i - 1], 1)] = list[0][i - 1][0]  # 首行插入参数
        for j in range(1, len(list) + 1):
            ws['%s%s' % (letters[i - 1], j + 1)] = list[j - 1][i - 1][1] #其余行插入数据
    r.save(file_path)
excel_ins('F:\excel\ok.xlsx','ok',norm_list(data_a))


